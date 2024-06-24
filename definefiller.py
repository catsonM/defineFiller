import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def complete_modules_fds(df):
    df.iloc[:, 0] = df.iloc[:, 0].ffill()  # Fill missing modules
    df.iloc[:, 1] = df.iloc[:, 1].ffill()  # Fill missing FDs
    return df

def complete_define_values(df):
    acp_columns = df.iloc[3, 3:].dropna().reset_index(drop=True)
    if 'ACP29036' in acp_columns.values:
        default_acp_index = acp_columns[acp_columns == 'ACP29036'].index[0] + 3
        default_values = df.iloc[6:, default_acp_index]
        # Fill missing define values
        for col in range(3, df.shape[1]):
            if col == default_acp_index:  # Skip the default ACP column
                continue
            df.iloc[6:, col] = df.iloc[6:, col].fillna(default_values)
    else:
        messagebox.showerror("Processing Error", "ACP29036 not found in the data.")
        return None
    return df

def process_raw_data(file_path):
    raw_df = pd.read_excel(file_path, header=None)
    
    # Fill missing modules and FDs
    raw_df = complete_modules_fds(raw_df)
    
    # Fill missing define values
    raw_df = complete_define_values(raw_df)
    
    return raw_df

def copy_styles(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
            if cell.hyperlink:
                new_cell.hyperlink = cell.hyperlink
            if cell.comment:
                new_cell.comment = cell.comment

    for merged_cell in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_cell))

def complete_files():
    input_path = entry_input.get()
    if not input_path:
        messagebox.showerror("Input Error", "Please select a raw file.")
        return

    pi_no = entry_pi_no.get()
    beta_version = entry_beta_version.get()
    if not pi_no or not beta_version:
        proceed = messagebox.askyesno("Missing Information", "PI no. or Beta Version is missing. Do you want to proceed?")
        if not proceed:
            return

    output_dir = entry_output.get()
    if not output_dir:
        output_dir = os.path.dirname(input_path)

    if subfolder_var.get():
        subfolder_path = os.path.join(output_dir, 'completed_files')
        os.makedirs(subfolder_path, exist_ok=True)
        output_dir = subfolder_path

    input_filename = os.path.basename(input_path)
    input_name, input_ext = os.path.splitext(input_filename)
    output_filename = f"afill_{input_name}"
    if pi_no:
        output_filename += f"_{pi_no}"
    if beta_version:
        output_filename += f"_{beta_version}"
    output_filename += input_ext
    output_file = os.path.join(output_dir, output_filename)

    if os.path.exists(output_file):
        overwrite = messagebox.askyesno("Overwrite Confirmation", f"{output_file} already exists. Do you want to overwrite it?")
        if not overwrite:
            return

    try:
        completed_df = process_raw_data(input_path)
        if completed_df is None:
            return

        # Load the workbook and worksheet to preserve styles
        wb = load_workbook(input_path)
        ws = wb.active

        # Save the completed data to a new workbook
        completed_df.to_excel(output_file, index=False, header=False)

        # Load the new workbook and copy styles
        completed_wb = load_workbook(output_file)
        completed_ws = completed_wb.active

        copy_styles(ws, completed_ws)

        # Save the completed workbook with styles
        completed_wb.save(output_file)

        messagebox.showinfo("Success", f"File completed and saved to: {output_file}")
    except Exception as e:
        import traceback
        error_message = traceback.format_exc()
        print("An error occurred:", error_message)  # Debugging line
        messagebox.showerror("Processing Error", f"An error occurred: {error_message}")

def select_input_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry_input.delete(0, tk.END)
    entry_input.insert(0, file_path)

def select_output_folder():
    folder_path = filedialog.askdirectory()
    entry_output.delete(0, tk.END)
    entry_output.insert(0, folder_path)

def show_help():
    help_text = (
        "DefineFiller, converts raw Excel files to completed ones with the specified format.\n\n"
        "Steps to use the program:\n"
        "1. Select the raw Excel file by clicking the 'Browse' button.\n"
        "2. Enter the PI no. and Beta Version in the respective fields. These fields are optional, but you will be prompted if they are left empty.\n"
        "3. Optionally, select an output folder where the completed file will be saved.\n"
        "4. If you want to save the file in a subfolder, check the 'Save in subfolder' option.\n"
        "5. Click the 'Complete' button to start the conversion process.\n\n"
        "Note: The output file will have a prefix 'afill_' to indicate that it was autofilled by DefineFiller.\n\n"
        "Author: Mert Can Catoglu\n"
        "E-posta: mertcan.catoglu@tr.bosch.com"
    )
    messagebox.showinfo("Help", help_text)

app = tk.Tk()
app.title("DefineFiller")

frame = tk.Frame(app)
frame.pack(padx=10, pady=10)

label_input = tk.Label(frame, text="Select Raw File:")
label_input.grid(row=0, column=0, sticky=tk.W, pady=5)
entry_input = tk.Entry(frame, width=50)
entry_input.grid(row=0, column=1, pady=5)
button_browse = tk.Button(frame, text="Browse", command=select_input_file)
button_browse.grid(row=0, column=2, padx=5, pady=5)

label_pi_no = tk.Label(frame, text="PI no.:")
label_pi_no.grid(row=1, column=0, sticky=tk.W, pady=5)
entry_pi_no = tk.Entry(frame, width=20)
entry_pi_no.grid(row=1, column=1, sticky=tk.W, pady=5)

label_beta_version = tk.Label(frame, text="Beta Version:")
label_beta_version.grid(row=2, column=0, sticky=tk.W, pady=5)
entry_beta_version = tk.Entry(frame, width=20)
entry_beta_version.grid(row=2, column=1, sticky=tk.W, pady=5)

label_output = tk.Label(frame, text="Select Output Folder (Optional):")
label_output.grid(row=3, column=0, sticky=tk.W, pady=5)
entry_output = tk.Entry(frame, width=50)
entry_output.grid(row=3, column=1, pady=5)
button_browse_output = tk.Button(frame, text="Browse", command=select_output_folder)
button_browse_output.grid(row=3, column=2, padx=5, pady=5)

subfolder_var = tk.BooleanVar()
checkbox_subfolder = tk.Checkbutton(frame, text="Save in subfolder", variable=subfolder_var)
checkbox_subfolder.grid(row=4, column=1, sticky=tk.W, pady=5)

button_complete = tk.Button(frame, text="Complete", command=complete_files)
button_complete.grid(row=5, column=1, pady=10)
button_complete.bind("<Enter>", lambda e: button_complete.config(bg="lightblue"))
button_complete.bind("<Leave>", lambda e: button_complete.config(bg="SystemButtonFace"))

button_help = tk.Button(frame, text="Help", command=show_help)
button_help.grid(row=5, column=2, pady=10)
button_help.bind("<Enter>", lambda e: button_help.config(bg="lightblue"))
button_help.bind("<Leave>", lambda e: button_help.config(bg="SystemButtonFace"))

app.mainloop()
